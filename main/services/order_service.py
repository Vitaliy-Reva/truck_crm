from openpyxl import load_workbook
from ..models import Order, Client

class OrderService:
    @staticmethod
    def create_order(data: dict, client):
        data['client_id'] = client
        filename = data.get("payment")
        order = Order.objects.create(**data)
        workbook = load_workbook(filename=filename)
        print(workbook)
        client.save()
        return order
    
    def update_order(data: dict, order: Order, client: Client, filename: str):
        data['client_id'] = client

        workbook = load_workbook(filename=f'./media/payments/{filename}')
        worksheet = workbook['Sheet1']
        for cell in worksheet.values:
            print(cell)

        order.order_name = data.get("order_name", order.order_name)
        order.price = data.get("price", order.price)
        order.payment = data.get("payment", order.payment)
        order.date = data.get("date", order.date)

        order.save()
        client.save()
        return order